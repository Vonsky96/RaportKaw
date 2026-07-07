from openpyxl import Workbook

def generuj_raport(plik):

    wb = Workbook()
    ws = wb.active

    ws.title = "Raport"

    ws["A1"] = "Program działa!"
    ws["A2"] = f"Wczytany plik: {plik}"

    wb.save("Raport_Kawy.xlsx")